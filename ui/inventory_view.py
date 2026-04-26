from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem,
    QDialog, QFormLayout, QLineEdit, QDateEdit,
    QMessageBox, QHeaderView
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor
from models import Producto
from database import get_session
from datetime import datetime

class InventoryView(QWidget):

    def __init__(self, engine):
        super().__init__()
        self.engine  = engine
        self.session = get_session()
        self._construir_ui()
        self.cargar_productos()

    def _construir_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(15)

        # Título
        titulo = QLabel("📦  Inventario de Productos")
        titulo.setStyleSheet("""
            font-size: 22px;
            font-weight: bold;
            color: #1E40AF;
        """)
        layout.addWidget(titulo)

        # ── Botones superiores ─────────────────────────────────────────────
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        btn_agregar = QPushButton("➕  Agregar")
        btn_agregar.setFixedHeight(38)
        btn_agregar.setStyleSheet("""
            QPushButton {
                background-color: #2563EB;
                color: white;
                padding: 0px 18px;
                border-radius: 7px;
                font-size: 13px;
                font-weight: bold;
                border: none;
            }
            QPushButton:hover { background-color: #1D4ED8; }
        """)
        btn_agregar.clicked.connect(self._abrir_formulario)

        btn_refrescar = QPushButton("🔄  Refrescar")
        btn_refrescar.setFixedHeight(38)
        btn_refrescar.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #374151;
                padding: 0px 18px;
                border-radius: 7px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #D1D5DB;
            }
            QPushButton:hover { background-color: #F9FAFB; }
        """)
        btn_refrescar.clicked.connect(self.cargar_productos)

        btn_excel = QPushButton("📊  Excel")
        btn_excel.setFixedHeight(38)
        btn_excel.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #166534;
                padding: 0px 18px;
                border-radius: 7px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #BBF7D0;
            }
            QPushButton:hover { background-color: #F0FDF4; }
        """)
        btn_excel.clicked.connect(self._exportar_excel)

        btn_pdf = QPushButton("📄  PDF")
        btn_pdf.setFixedHeight(38)
        btn_pdf.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #991B1B;
                padding: 0px 18px;
                border-radius: 7px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #FECACA;
            }
            QPushButton:hover { background-color: #FFF5F5; }
        """)
        btn_pdf.clicked.connect(self._exportar_pdf)

        btn_layout.addWidget(btn_agregar)
        btn_layout.addWidget(btn_refrescar)
        btn_layout.addWidget(btn_excel)
        btn_layout.addWidget(btn_pdf)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # ── Tabla ──────────────────────────────────────────────────────────
        self.tabla = QTableWidget()
        self.tabla.setColumnCount(8)
        self.tabla.setHorizontalHeaderLabels([
            "ID", "Nombre", "Categoría", "Precio Base",
            "Stock", "Stock Máximo", "Vencimiento", "Riesgo"
        ])
        self.tabla.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.verticalHeader().setVisible(False)
        self.tabla.setShowGrid(False)
        self.tabla.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;
                color: #111827;
                border-radius: 10px;
                border: 1px solid #E5E7EB;
                font-size: 13px;
                gridline-color: #F3F4F6;
            }
            QTableWidget::item {
                padding: 10px 8px;
                color: #374151;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #DBEAFE;
                color: #1D4ED8;
            }
            QTableWidget::item:alternate {
                background-color: #F9FAFB;
            }
            QHeaderView::section {
                background-color: #F3F4F6;
                color: #6B7280;
                padding: 10px 8px;
                font-size: 12px;
                font-weight: bold;
                border: none;
                border-bottom: 2px solid #E5E7EB;
            }
        """)
        layout.addWidget(self.tabla)

        # ── Botones inferiores ─────────────────────────────────────────────
        btn_layout2 = QHBoxLayout()
        btn_layout2.setSpacing(10)

        btn_editar = QPushButton("✏️  Editar")
        btn_editar.setFixedHeight(38)
        btn_editar.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #92400E;
                padding: 0px 18px;
                border-radius: 7px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #FDE68A;
            }
            QPushButton:hover { background-color: #FFFBEB; }
        """)
        btn_editar.clicked.connect(self._editar_producto)

        btn_eliminar = QPushButton("🗑️  Eliminar")
        btn_eliminar.setFixedHeight(38)
        btn_eliminar.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #991B1B;
                padding: 0px 18px;
                border-radius: 7px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #FECACA;
            }
            QPushButton:hover { background-color: #FEF2F2; }
        """)
        btn_eliminar.clicked.connect(self._eliminar_producto)

        btn_layout2.addWidget(btn_editar)
        btn_layout2.addWidget(btn_eliminar)
        btn_layout2.addStretch()
        layout.addLayout(btn_layout2)

    # ── Cargar productos ───────────────────────────────────────────────────
    def cargar_productos(self):
        self.session.expire_all()
        productos = self.session.query(Producto).filter(Producto.activo == 1).all()
        self.tabla.setRowCount(len(productos))

        badges = {
            "critico": ("Crítico", "#991B1B", "#FECACA"),
            "alto":    ("Alto",    "#9A3412", "#FED7AA"),
            "medio":   ("Medio",   "#854D0E", "#FEF08A"),
            "normal":  ("Normal",  "#166534", "#DCFCE7"),
        }

        for fila, p in enumerate(productos):
            vencimiento = p.fecha_vencimiento.strftime("%Y-%m-%d") \
                          if p.fecha_vencimiento else "Sin vencimiento"

            valores = [
                str(p.id), p.nombre, p.categoria,
                f"${p.precio_base:,.0f}", str(p.stock),
                str(p.stock_maximo), vencimiento,
                p.nivel_riesgo.upper()
            ]

            etiqueta, color_txt, color_bg = badges.get(
                p.nivel_riesgo, ("Normal", "#166534", "#DCFCE7")
            )

            for col, valor in enumerate(valores):
                item = QTableWidgetItem(valor)
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                if col == 7:
                    item.setForeground(QColor(color_txt))
                    item.setBackground(QColor(color_bg))
                    item.setText(etiqueta)
                self.tabla.setItem(fila, col, item)

        self.tabla.setRowHeight(0, 42)

    # ── Abrir formulario ───────────────────────────────────────────────────
    def _abrir_formulario(self, producto=None):
        dialogo = ProductoDialog(producto, self)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            datos = dialogo.obtener_datos()
            if producto:
                producto.nombre            = datos["nombre"]
                producto.categoria         = datos["categoria"]
                producto.precio_base       = datos["precio_base"]
                producto.stock             = datos["stock"]
                producto.stock_maximo      = datos["stock_maximo"]
                producto.fecha_vencimiento = datos["fecha_vencimiento"]
            else:
                self.session.add(Producto(**datos))
            self.session.commit()
            self.cargar_productos()

    # ── Editar ─────────────────────────────────────────────────────────────
    def _editar_producto(self):
        fila = self.tabla.currentRow()
        if fila < 0:
            QMessageBox.warning(self, "Aviso", "Selecciona un producto primero.")
            return
        producto_id = int(self.tabla.item(fila, 0).text())
        producto    = self.session.get(Producto, producto_id)
        self._abrir_formulario(producto)

    # ── Eliminar ───────────────────────────────────────────────────────────
    def _eliminar_producto(self):
        fila = self.tabla.currentRow()
        if fila < 0:
            QMessageBox.warning(self, "Aviso", "Selecciona un producto primero.")
            return
        confirmacion = QMessageBox.question(
            self, "Confirmar",
            "¿Estás seguro de que deseas eliminar este producto?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if confirmacion == QMessageBox.StandardButton.Yes:
            producto_id = int(self.tabla.item(fila, 0).text())
            producto    = self.session.get(Producto, producto_id)
            producto.activo = 0
            self.session.commit()
            self.cargar_productos()

    # ── Exportar Excel ─────────────────────────────────────────────────────
    def _exportar_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from config import REPORTS_DIR, CONFIG

        productos = self.session.query(Producto).filter(Producto.activo == 1).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        encabezados = ["ID", "Nombre", "Categoría", "Precio Base",
                       "Stock", "Stock Máximo", "Vencimiento", "Riesgo"]
        for col, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=enc)
            celda.font      = Font(bold=True, color="FFFFFF")
            celda.fill      = PatternFill("solid", fgColor="1E40AF")
            celda.alignment = Alignment(horizontal="center")

        colores = {
            "critico": "FECACA",
            "alto":    "FED7AA",
            "medio":   "FEF08A",
            "normal":  "DCFCE7",
        }

        for fila, p in enumerate(productos, 2):
            vencimiento = p.fecha_vencimiento.strftime("%Y-%m-%d") \
                          if p.fecha_vencimiento else "Sin vencimiento"
            valores = [p.id, p.nombre, p.categoria, p.precio_base,
                       p.stock, p.stock_maximo, vencimiento,
                       p.nivel_riesgo.upper()]
            color = colores.get(p.nivel_riesgo, "FFFFFF")
            for col, valor in enumerate(valores, 1):
                celda = ws.cell(row=fila, column=col, value=valor)
                celda.fill = PatternFill("solid", fgColor=color)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        ruta = REPORTS_DIR / f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(ruta)
        QMessageBox.information(self, "✅ Exportado", f"Excel guardado en:\n{ruta}")

    # ── Exportar PDF ───────────────────────────────────────────────────────
    def _exportar_pdf(self):
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet
        from config import REPORTS_DIR, CONFIG

        productos = self.session.query(Producto).filter(Producto.activo == 1).all()
        ruta      = REPORTS_DIR / f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        doc      = SimpleDocTemplate(str(ruta), pagesize=landscape(letter))
        estilos  = getSampleStyleSheet()
        elementos = []

        titulo = Paragraph(
            f"<b>Reporte de Inventario — {CONFIG.get('nombre_negocio', 'Mi Tienda')}</b>",
            estilos["Title"]
        )
        fecha = Paragraph(
            f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            estilos["Normal"]
        )
        elementos += [titulo, fecha, Spacer(1, 20)]

        datos = [["ID", "Nombre", "Categoría", "Precio",
                  "Stock", "Stock Máx.", "Vencimiento", "Riesgo"]]

        colores_pdf = {
            "critico": colors.HexColor("#FECACA"),
            "alto":    colors.HexColor("#FED7AA"),
            "medio":   colors.HexColor("#FEF08A"),
            "normal":  colors.white,
        }

        filas_colores = []
        for i, p in enumerate(productos, 1):
            vencimiento = p.fecha_vencimiento.strftime("%Y-%m-%d") \
                          if p.fecha_vencimiento else "Sin vencimiento"
            datos.append([
                str(p.id), p.nombre, p.categoria,
                f"${p.precio_base:,.0f}", str(p.stock),
                str(p.stock_maximo), vencimiento,
                p.nivel_riesgo.upper()
            ])
            filas_colores.append((i, colores_pdf.get(p.nivel_riesgo, colors.white)))

        tabla = Table(datos, repeatRows=1)
        estilo = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white]),
        ])
        for fila_idx, color in filas_colores:
            estilo.add("BACKGROUND", (0, fila_idx), (-1, fila_idx), color)

        tabla.setStyle(estilo)
        elementos.append(tabla)
        doc.build(elementos)

        QMessageBox.information(self, "✅ Exportado", f"PDF guardado en:\n{ruta}")


# ── Diálogo agregar / editar ───────────────────────────────────────────────
class ProductoDialog(QDialog):

    def __init__(self, producto=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Agregar Producto" if not producto else "Editar Producto")
        self.setMinimumWidth(420)
        self._construir_ui(producto)
        self.setStyleSheet("""
            QDialog {
                background-color: #FFFFFF;
            }
            QLabel {
                color: #374151;
                font-size: 13px;
                font-weight: bold;
            }
            QLineEdit, QDateEdit {
                background-color: #F9FAFB;
                color: #111827;
                border: 1px solid #D1D5DB;
                border-radius: 7px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QLineEdit:focus, QDateEdit:focus {
                border: 1.5px solid #3B82F6;
                background-color: #FFFFFF;
            }
        """)

    def _construir_ui(self, producto):
        layout = QFormLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(25, 25, 25, 25)

        self.nombre    = QLineEdit(producto.nombre    if producto else "")
        self.categoria = QLineEdit(producto.categoria if producto else "")
        self.precio    = QLineEdit(str(producto.precio_base)  if producto else "")
        self.stock     = QLineEdit(str(producto.stock)        if producto else "")
        self.stock_max = QLineEdit(str(producto.stock_maximo) if producto else "")

        self.vencimiento = QDateEdit()
        self.vencimiento.setCalendarPopup(True)
        if producto and producto.fecha_vencimiento:
            fecha = producto.fecha_vencimiento
            self.vencimiento.setDate(QDate(fecha.year, fecha.month, fecha.day))
        else:
            self.vencimiento.setDate(QDate.currentDate())

        layout.addRow("Nombre:",       self.nombre)
        layout.addRow("Categoría:",    self.categoria)
        layout.addRow("Precio Base:",  self.precio)
        layout.addRow("Stock:",        self.stock)
        layout.addRow("Stock Máximo:", self.stock_max)
        layout.addRow("Vencimiento:",  self.vencimiento)

        btn_layout   = QHBoxLayout()
        btn_guardar  = QPushButton("💾  Guardar")
        btn_cancelar = QPushButton("Cancelar")
        btn_guardar.setFixedHeight(40)
        btn_cancelar.setFixedHeight(40)
        btn_guardar.clicked.connect(self.accept)
        btn_cancelar.clicked.connect(self.reject)
        btn_guardar.setStyleSheet("""
            QPushButton {
                background-color: #2563EB;
                color: white;
                padding: 0px 20px;
                border-radius: 7px;
                font-weight: bold;
                border: none;
            }
            QPushButton:hover { background-color: #1D4ED8; }
        """)
        btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #6B7280;
                padding: 0px 20px;
                border-radius: 7px;
                font-weight: bold;
                border: 1px solid #D1D5DB;
            }
            QPushButton:hover { background-color: #F9FAFB; }
        """)
        btn_layout.addWidget(btn_guardar)
        btn_layout.addWidget(btn_cancelar)
        layout.addRow(btn_layout)

    def obtener_datos(self) -> dict:
        fecha_q = self.vencimiento.date()
        return {
            "nombre"           : self.nombre.text(),
            "categoria"        : self.categoria.text(),
            "precio_base"      : float(self.precio.text()),
            "stock"            : int(self.stock.text()),
            "stock_maximo"     : int(self.stock_max.text()),
            "fecha_vencimiento": datetime(fecha_q.year(), fecha_q.month(), fecha_q.day()),
        }